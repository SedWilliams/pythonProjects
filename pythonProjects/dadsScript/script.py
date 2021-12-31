from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

spreadsheet = input("Name of file (Including .xlsx): ")
wb = load_workbook(spreadsheet)
ws = wb.active
days_worked = 0
hours_worked = 0


minrow = int(input("What row do the dates start from?: "))
maxrow = int(input("What row do the dates end?: "))
mincol = int(input("What column do the dates start at?: "))
maxcol = int(input("What column do the dates end on?: "))
output_col = maxcol + 1

for row in ws.iter_rows(min_row=minrow, min_col=mincol, max_col=maxcol, max_row=maxrow, values_only=True):

    shift_type = input("Part time or Full time? (p/f): ")
    if shift_type == "p":
        hours_worked = days_worked * 4
    elif shift_type == "f":
        hours_worked = days_worked * 8

    days_worked = row[1] - row[0]
    print(days_worked)

    print(hours_worked, "<-- HOURS")

#NOT WORKING
#SUPPOSED TO ITERATE THROUGH HOURS WORKED COLUMN AND UPDATE VALUES IN COLUMN TO HOURS_WORKED
#FIXED, TUPLES ARE IMMUTABLE SO HAD TO CHANGE IT TO A LIST
for rows in ws.iter_rows(min_row=minrow, min_col=output_col, max_row=maxrow, values_only=True):
    rowsList = list(rows)
    rowsList[0] = hours_worked

#NOT WORKING
#WILL NOT SAVE TO SPREADSHEET
wb.save(spreadsheet)
